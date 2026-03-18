import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from dataclasses import dataclass

@dataclass
class RowData:
    index: int
    sheet: str
    tcode: str
    explanation: str
    params: str
    mode: str

class ExcelManager:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = None
        self.active_sheet = None

    def __enter__(self):
        self.wb = openpyxl.load_workbook(self.file_path)
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.wb.save(self.file_path)

    def get_sheets(self):
        return self.wb.sheetnames

    def read_rows(self, sheet_name: str):
        ws = self.wb[sheet_name]
        rows = []
        header = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
        
        def find_col(names):
            for name in names:
                if name in header: return header.index(name)
            return None

        idx_tcode = find_col(["tcode", "transação", "transacao", "transaction"])
        idx_param = find_col(["parâmetro", "parametro", "parameters", "parameter"])
        idx_expl = find_col(["explanation", "test explanation", "descrição", "descricao"])
        idx_mode = find_col(["modo", "mode"])

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            tcode = str(row[idx_tcode]).strip() if idx_tcode is not None else ""
            if not tcode or tcode == "None": continue
            
            rows.append(RowData(
                index=i, 
                sheet=sheet_name, 
                tcode=tcode,
                explanation=str(row[idx_expl]) if idx_expl is not None else "", 
                params=str(row[idx_param]) if idx_param is not None else "", 
                mode=str(row[idx_mode]) if idx_mode is not None else "executar"
            ))
        return rows

    def write_result(self, sheet_name, row_idx, result):
        ws = self.wb[sheet_name]
        ws.cell(row_idx, 7).value = result.status
        ws.cell(row_idx, 8).value = result.message
        if result.evidence_path:
            cell = ws.cell(row_idx, 9)
            cell.value = "Ver Evidência"
            cell.hyperlink = result.evidence_path