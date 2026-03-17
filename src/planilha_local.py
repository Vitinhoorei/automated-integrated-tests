from __future__ import annotations

from dataclasses import dataclass
from typing import List, Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# MODELO DE LINHA
@dataclass
class SheetRow:
    sheet_name: str
    row_index: int
    scenario: str
    tcode: str
    explanation: str
    parameter: str
    mode: str


# UTILITÁRIOS
def _norm(value: str) -> str:
    return str(value or "").strip().lower()


def _find_col(header_row, *names: str) -> Optional[int]:
    wanted = {_norm(n) for n in names if n}
    for idx, cell in enumerate(header_row, start=1):
        if _norm(cell.value) in wanted:
            return idx
    return None


def _normalize_mode(value: str) -> str:
    mode = _norm(value)

    if not mode:
        return ""

    aliases = {
        "executar": "executar",
        "exec": "executar",
        "real": "executar",

        "simulado": "simulado",
        "simulada": "simulado",
        "simulacao": "simulado",
        "simulação": "simulado",
        "simulate": "simulado",
        "teste": "simulado",
    }

    return aliases.get(mode, mode)


# LEITURA
def read_rows(xlsx_path: str, sheet_name: str) -> List[SheetRow]:
    wb = load_workbook(xlsx_path)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba '{sheet_name}' não existe.")

    ws = wb[sheet_name]
    header = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]

    col_scenario = _find_col(header, "Scenario", "Cenario", "Cenário")
    col_scen_expl = _find_col(header, "Scenario Explanation")
    col_tcode = _find_col(header, "Transação", "Transacao", "TCODE", "Transaction")
    col_test_expl = _find_col(header, "Test Explanation", "Explanation", "Test", "Descricao")
    col_param = _find_col(header, "Parâmetro", "Parametro", "Parameters", "Parameter")
    col_mode = _find_col(header, "Modo", "Mode")
    rows: List[SheetRow] =[]

    for r in range(2, ws.max_row + 1):
        tcode_raw = ws.cell(r, col_tcode).value if col_tcode else ""
        tcode = str(tcode_raw or "").strip()
        if not tcode:
            continue

        scenario = str(ws.cell(r, col_scenario).value or "").strip() if col_scenario else ""
        scen_expl = str(ws.cell(r, col_scen_expl).value or "").strip() if col_scen_expl else ""
        test_expl = str(ws.cell(r, col_test_expl).value or "").strip() if col_test_expl else ""

        textos = [txt for txt in[scen_expl, test_expl] if txt]
        explanation = " - ".join(textos)

        parameter = str(ws.cell(r, col_param).value or "").strip() if col_param else ""

        raw_mode = ws.cell(r, col_mode).value if col_mode else ""
        mode = _normalize_mode(raw_mode)

        rows.append(
            SheetRow(
                sheet_name,
                r,
                scenario,
                tcode,
                explanation,
                parameter,
                mode
            )
        )

    return rows

def list_sheet_names(xlsx_path: str) -> List[str]:
    return list(load_workbook(xlsx_path).sheetnames)


# COLUNAS DE STATUS
def ensure_status_columns(xlsx_path: str, sheet_name: str) -> dict:
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]
    header = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]
    col_status = _find_col(header, "Status")

    if not col_status:
        raise ValueError("Coluna 'Status' não encontrada.")

    col_source = col_status + 1
    col_message = col_status + 2
    col_suggestion = col_status + 3
    col_fix_conf = col_status + 4
    col_fix_just = col_status + 5
    col_evidence = col_status + 6

    headers = {
        col_source: "Status Source",
        col_message: "Status Message",
        col_suggestion: "Suggested Fix",
        col_fix_conf: "Fix Confidence",
        col_fix_just: "Fix Justification",
        col_evidence: "Evidence Path"
    }

    for col, title in headers.items():
        if not (ws.cell(1, col).value or "").strip():
            ws.cell(1, col).value = title

    wb.save(xlsx_path)

    return {
        "status": col_status,
        "source": col_source,
        "message": col_message,
        "suggestion": col_suggestion,
        "fix_confidence": col_fix_conf,
        "fix_justification": col_fix_just,
        "evidence": col_evidence,
    }


# ESCRITA (LEGADO - Mantido para compatibilidade com a main)
def write_status_triplet(
    xlsx_path: str,
    sheet_name: str,
    row_index: int,
    status: str,
    source: str,
    message: str,
    evidence_path: str | None = None,
    suggestion: str = "",
) -> None:
    cols = ensure_status_columns(xlsx_path, sheet_name)
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]

    ws.cell(row_index, cols["status"]).value = status
    ws.cell(row_index, cols["source"]).value = source
    ws.cell(row_index, cols["message"]).value = message

    ws.cell(row_index, cols["message"]).alignment = Alignment(wrap_text=True)
    ws.cell(row_index, cols["source"]).alignment = Alignment(wrap_text=True)

    ws.cell(row_index, cols["suggestion"]).value = suggestion
    ws.cell(row_index, cols["suggestion"]).alignment = Alignment(wrap_text=True)

    if evidence_path:
        ev_cell = ws.cell(row_index, cols["evidence"])
        ev_cell.value = evidence_path
        ev_cell.hyperlink = evidence_path
        ev_cell.style = "Hyperlink"

    wb.save(xlsx_path)


# ESCRITA (COMPLETA - NOVA)
def write_status_with_fix_details(
    xlsx_path: str,
    sheet_name: str,
    row_index: int,
    status: str,
    source: str,
    message: str,
    suggested_fix: str,
    fix_confidence: int,
    fix_justification: str,
    evidence_path: str | None = None,
):
    cols = ensure_status_columns(xlsx_path, sheet_name)
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]

    ws.cell(row_index, cols["status"]).value = status
    ws.cell(row_index, cols["source"]).value = source
    ws.cell(row_index, cols["message"]).value = message
    ws.cell(row_index, cols["suggestion"]).value = suggested_fix
    ws.cell(row_index, cols["fix_confidence"]).value = fix_confidence
    ws.cell(row_index, cols["fix_justification"]).value = fix_justification
    align = Alignment(vertical="top", wrap_text=True)
    ws.cell(row_index, cols["message"]).alignment = align
    ws.cell(row_index, cols["suggestion"]).alignment = align
    ws.cell(row_index, cols["fix_justification"]).alignment = align

    if evidence_path:
        ev_cell = ws.cell(row_index, cols["evidence"])
        ev_cell.value = evidence_path
        ev_cell.hyperlink = evidence_path
        ev_cell.style = "Hyperlink"

    wb.save(xlsx_path)


# AUTO AJUSTE DE COLUNA
def _auto_adjust_columns(ws):
    for column in ws.columns:
        max_length = 0
        column = list(column)
        letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted = min(max_length + 3, 80)
        ws.column_dimensions[letter].width = adjusted


# FORMATAÇÃO DA PLANILHA
def format_output_sheet(xlsx_path: str, sheet_name: str) -> None:
    cols = ensure_status_columns(xlsx_path, sheet_name)
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]
    max_row = ws.max_row
    max_col = ws.max_column

    header_fill = PatternFill("solid", fgColor="00579F")
    header_font = Font(color="FFFFFF", bold=True)

    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")

    zebra_fill = PatternFill("solid", fgColor="F7F9FB")

    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    # CABEÇALHO
    for c in range(1, max_col + 1):
        cell = ws.cell(1, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=False
        )

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    # LINHAS
    for r in range(2, max_row + 1):
        status = (ws.cell(r, cols["status"]).value or "").upper()

        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if r % 2 == 0:
                cell.fill = zebra_fill

        if status == "PASS":
            ws.cell(r, cols["status"]).fill = pass_fill

        if status == "FAIL":
            for c in range(1, max_col + 1):
                ws.cell(r, c).fill = fail_fill

    # FILTRO
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    # AUTO AJUSTE
    _auto_adjust_columns(ws)

    wb.save(xlsx_path)