from __future__ import annotations

from dataclasses import dataclass

from openpyxl import load_workbook

from sap_core_next.core.models import ExecutionReport


@dataclass
class OpenpyxlResultSink:
    xlsx_path: str
    sheet_name: str

    def write_report(self, report: ExecutionReport) -> None:
        wb = load_workbook(self.xlsx_path)
        ws = wb[self.sheet_name]

        # cria headers mínimos no final
        start_col = ws.max_column + 1
        headers = ["Engine Status", "Engine Source", "Engine Message", "Engine Evidence", "Engine Attempts"]
        for i, h in enumerate(headers):
            ws.cell(1, start_col + i).value = h

        # index por row da metadata
        row_index = {rec.step.metadata.get("row"): rec for rec in report.records}
        for r, rec in row_index.items():
            if not r:
                continue
            ws.cell(r, start_col).value = rec.status.value
            ws.cell(r, start_col + 1).value = rec.source
            ws.cell(r, start_col + 2).value = rec.message
            ws.cell(r, start_col + 3).value = rec.evidence_path
            ws.cell(r, start_col + 4).value = len(rec.attempts)

        wb.save(self.xlsx_path)
