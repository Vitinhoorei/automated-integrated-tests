from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, List

from openpyxl import load_workbook

from sap_core_next.core.models import ScenarioStep


def _norm(v: str) -> str:
    return str(v or "").strip().lower()


def _find_col(header, *names: str) -> int | None:
    wanted = {_norm(n) for n in names}
    for idx, cell in enumerate(header, start=1):
        if _norm(cell.value) in wanted:
            return idx
    return None


def _parse_parameters(raw: str) -> Dict[str, str]:
    parts = [p.strip() for p in str(raw or "").split("|") if p.strip()]
    out: Dict[str, str] = {}
    for p in parts:
        if ":" not in p:
            continue
        k, v = p.split(":", 1)
        out[k.strip()] = v.strip()
    return out


@dataclass
class OpenpyxlScenarioSource:
    xlsx_path: str
    sheet_name: str

    def read_steps(self) -> List[ScenarioStep]:
        wb = load_workbook(self.xlsx_path)
        ws = wb[self.sheet_name]
        header = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]

        c_module = _find_col(header, "Module", "Módulo", "Modulo")
        c_cmd = _find_col(header, "Command", "TCODE", "Transação", "Transacao")
        c_expl = _find_col(header, "Explanation", "Test Explanation", "Descricao")
        c_params = _find_col(header, "Parameters", "Parâmetro", "Parametro")
        c_mode = _find_col(header, "Mode", "Modo")

        steps: List[ScenarioStep] = []
        for r in range(2, ws.max_row + 1):
            cmd = str(ws.cell(r, c_cmd).value or "").strip() if c_cmd else ""
            if not cmd:
                continue

            module = str(ws.cell(r, c_module).value or "").strip() if c_module else "PM"
            explanation = str(ws.cell(r, c_expl).value or "").strip() if c_expl else ""
            params_raw = str(ws.cell(r, c_params).value or "").strip() if c_params else ""
            mode = str(ws.cell(r, c_mode).value or "real").strip() if c_mode else "real"

            steps.append(
                ScenarioStep(
                    step_id=f"{self.sheet_name}_r{r}",
                    module=module,
                    command=cmd,
                    explanation=explanation,
                    parameters=_parse_parameters(params_raw),
                    mode=mode,
                    metadata={"sheet": self.sheet_name, "row": r, "xlsx_path": self.xlsx_path},
                )
            )
        return steps
